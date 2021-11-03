import openpyxl
from datetime import date
from openpyxl.styles import Font
from openpyxl.styles import colors
from pathlib import Path
from datetime import date, timedelta
from tqdm import trange
import forecast
import xlrd

xlsx_file = Path('../data', 'sample_excel.xlsx')
# print('file location: ', xlsx_file)
# wb_obj = openpyxl.load_workbook(xlsx_file)
#
# print(wb_obj.sheetnames)


# for row in sheet.iter_rows(max_row=6):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()


# test writing to excel file
def test_writing_file():
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj["t_mm_prognoze"]
    print(sheet)
    t = sheet['A1']
    t.value = 'Hello test'
    # add date at the end of filename
    today = date.today().strftime('%d%m%Y')
    filename = "RU05_laika_apstakli_fakts_prognoze_%s.xlsx" % today
    wb_obj.save(filename)


# Save new xlsx file and update dates in forecast sheet

def update_dates(sample_file_path):
    print('file location: ', sample_file_path)
    wb_obj = openpyxl.load_workbook(sample_file_path)
    sheet = wb_obj["t_mm_prognoze"]
    dd = 0
    for row in sheet.iter_rows(min_row=3, max_row=12, max_col=5, min_col=5):
        for cell in row:
            dd += 1
            val = (date.today() + timedelta(days=dd)).strftime('%d.%m.%Y')
            cell.value = val
            cell.font = Font(color="00000080", italic=True)  # for testing write in blue
            print(cell.value, end=" ")
    xlsx_filename = "RU05_laika_apstakli_fakts_prognoze_%s.xlsx" % date.today().strftime('%d%m%Y')
    wb_obj.save(xlsx_filename)
    return xlsx_filename


def update_forecast(filename):
    wb_obj = openpyxl.load_workbook(filename)
    # Read the forecast sheet t_mm_prognoze
    sheet = wb_obj["t_mm_prognoze"]

    dupl_st = ['Bauska', 'Gulbene', 'Kuldīga', 'Rūjiena', 'Stende']
    tripl_st = ['Zīlāni', 'Zosēni']
    station_dict = {'Ainaži': 'ainazi-47075', 'Alūksne': 'aluksne-4124', 'Bauska': 'bauska-4139', 'Daugavpils': 'daugavpils-4177', 'Dobele': 'dobele-4137', 'Gulbene': 'gulbene-4125', 'Jelgava': 'jelgava-4138', 'Kalnciems': 'kalnciems-322319', 'Kolka': 'kolka-47057', 'Kuldīga': 'kuldiga-4106', 'Lielpeči': 'ogre-47066', 'Liepāja': 'liepaja-4134', 'Mērsrags': 'mersrags-47065', 'Pāvilosta': 'pavilosta-47067', 'Piedruja': 'piedruja-322361', 'Rēzekne': 'rezekne-4140', 'Rīga': 'riga-4136', 'Rūjiena': 'rujiena-47073', 'Saldus': 'saldus-4135', 'Sigulda': 'sigulda-4103', 'Sīļi': 'silajani-322513', 'Skrīveri': 'skriveri-322563', 'Stende': 'stende-47083', 'Vičaki': 'ventspils-4123', 'Zīlāni': 'zilani-322369', 'Zosēni': 'vecpiebalga-322623'}

    listt, listmm = forecast.mm_temp_forecast_duplicates(station_dict, dupl_st, tripl_st)
    # l = 0           # iterable for temperature list

    for i in range(3, 353):
        l = i - 3
        cellref_t = sheet.cell(row=i, column=6)     # writing temp to f column
        cellref_t.value = listt[l]
        cellref_mm = sheet.cell(row=i, column=8)    # writing mm to H column
        cellref_mm.value = listmm[l]

    wb_obj.save(filename)


new_file = update_dates(xlsx_file)
update_forecast(new_file)

# write historical weather data from excel files to big excel



# station_path = Path('data', '20210512', 'abi_Ainaži05.05.2021.xls')
# station_wb = xlrd.open_workbook(station_path)
final_wb = openpyxl.load_workbook('RU05_laika_apstakli_fakts_prognoze_12052021.xlsx') # TODO add inside function as var
#
# # excel source data
# t_sheet = station_wb.sheet_by_name('1. tabula')
#
# # t_sheet = station_wb.get_sheet_by_name('1. tabula')
# mm_sheet = station_wb.sheet_by_name('2. tabula')
# # station_mm_sheet = station_wb['2. tabula']
#
# #excel destination data
#
final_t_sheet = final_wb['t_fakts']
final_mm_sheet = final_wb['mm_fakts']


# to destination excel file

# for i in range(3, 10):      # destination excel rows from - to
#     # get each row from source file into list
#     header = t_sheet.row_values(i-1, start_colx=0, end_colx=None)
#     print(header)
#     for j in range(4, 29):  # columns
#         # reading cell value from source excel file
#         c = header[j-4]
#         print('c value: ',c)
#         # writing the read value to destination excel file
#         final_t_sheet.cell(row=i, column=j).value = c
# final_wb.save('destination.xlsx')


# find station row in excel final file
def find_start_row(station):
    row_idx = False
    for row in final_t_sheet.iter_rows(min_col=2, max_col=2):   # B column in Excel file
        for cell in row:
            if cell.value == station:
                # print('Row number:', row[0].row)
                row_idx = row[0].row
    return row_idx


print(find_start_row('Zīlāni'))


# write temp history for one station in big excel file
# def write_t_history(src_filepath, dest_filepath, station):
#     station_wb = xlrd.open_workbook(src_filepath)
#     final_wb = openpyxl.load_workbook(dest_filepath)
#     t_ws = station_wb.sheet_by_name('1. tabula')
#     final_t_ws = final_wb['t_fakts']
#     # find station row?
#     st_row = find_start_row(station)
#     for i in range(st_row, st_row+7):  # destination excel rows from - to
#         # get each row from source file into list
#         header = t_ws.row_values(i - 1, start_colx=0, end_colx=None)
#         print(header)
#         for j in range(4, 29):  # columns
#             # reading cell value from source excel file
#             c = header[j - 4]
#             # writing the read value to destination excel file
#             final_t_ws.cell(row=i, column=j).value = c
#     final_wb.save('destination.xlsx')

dupl_st = ['Bauska', 'Gulbene', 'Kuldīga', 'Rūjiena', 'Stende']
tripl_st = ['Zīlāni', 'Zosēni']
station_dict = {'Ainaži': 'ainazi-47075', 'Alūksne': 'aluksne-4124', 'Bauska': 'bauska-4139', 'Daugavpils': 'daugavpils-4177', 'Dobele': 'dobele-4137', 'Gulbene': 'gulbene-4125', 'Jelgava': 'jelgava-4138', 'Kalnciems': 'kalnciems-322319', 'Kolka': 'kolka-47057', 'Kuldīga': 'kuldiga-4106', 'Lielpeči': 'ogre-47066', 'Liepāja': 'liepaja-4134', 'Mērsrags': 'mersrags-47065', 'Pāvilosta': 'pavilosta-47067', 'Piedruja': 'piedruja-322361', 'Rēzekne': 'rezekne-4140', 'Rīga': 'riga-4136', 'Rūjiena': 'rujiena-47073', 'Saldus': 'saldus-4135', 'Sigulda': 'sigulda-4103', 'Sīļi': 'silajani-322513', 'Skrīveri': 'skriveri-322563', 'Stende': 'stende-47083', 'Vičaki': 'ventspils-4123', 'Zīlāni': 'zilani-322369', 'Zosēni': 'vecpiebalga-322623'}


# write temp history for one station in big excel file
def write_t_history(src_filepath, dest_filepath, station):
    station_wb = xlrd.open_workbook(src_filepath)
    final_wb = openpyxl.load_workbook(dest_filepath)
    t_ws = station_wb.sheet_by_name('1. tabula')
    final_t_ws = final_wb['t_fakts']
    # find station row?
    str_row = find_start_row(station)   # TODO - wrap around try except - for cases when station not found in excel wb
    print(str_row)
    write_data_chunk(str_row, t_ws, final_t_ws)
    if station in dupl_st:
        write_data_chunk(str_row+7, t_ws, final_t_ws)
    elif station in tripl_st:
        write_data_chunk(str_row+7, t_ws, final_t_ws)
        write_data_chunk(str_row+14, t_ws, final_t_ws)
    final_wb.save(dest_filepath)


# write temp & rain history for one station in big excel file
def write_t_mm_history(src_filepath, dest_filepath, station):
    station_wb = xlrd.open_workbook(src_filepath)
    final_wb = openpyxl.load_workbook(dest_filepath)
    t_ws = station_wb.sheet_by_name('1. tabula')
    mm_ws = station_wb.sheet_by_name('2. tabula')
    final_t_ws = final_wb['t_fakts']
    final_mm_ws = final_wb['mm_fakts']
    # find station row?
    str_row = find_start_row(station)   # TODO - wrap around try except - for cases when station not found in excel wb
    print(str_row)
    # write temperature
    write_data_chunk(str_row, t_ws, final_t_ws)
    write_data_chunk(str_row, mm_ws, final_mm_ws)
    # if station in dupl_st:  # not needed with proper excel structure?
    #     write_data_chunk(str_row+7, t_ws, final_t_ws)    # double temp
    #     write_data_chunk(str_row+7, mm_ws, final_mm_ws)  # double rain
    # elif station in tripl_st:
    #     write_data_chunk(str_row+7, t_ws, final_t_ws)     # double temp
    #     write_data_chunk(str_row+14, t_ws, final_t_ws)
    #     write_data_chunk(str_row+7, mm_ws, final_mm_ws)   # double rain
    #     write_data_chunk(str_row+14, mm_ws, final_mm_ws)
    final_wb.save(dest_filepath)


def write_data_chunk(st_row, src_ws, dest_ws):
    h = 2
    for i in range(st_row, st_row+7):  # destination excel rows from - to
        # get each row from source file into list
        header = src_ws.row_values(h, start_colx=0, end_colx=None)
        print(header)
        h += 1
        for j in range(4, 29):  # columns
            # reading cell value from source excel file
            c = header[j - 4]
            # writing the read value to destination excel file
            dest_ws.cell(row=i, column=j).value = c


for key in station_dict:
    print(key)
    st_path = station_path = Path('../data', '20210512', 'abi_%s11.05.2021.xls' % key) #TODO automate folder name
    write_t_mm_history(st_path, '../RU05_laika_apstakli_fakts_prognoze_sample.xlsx', key)
