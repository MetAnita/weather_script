import os
import openpyxl
from pathlib import Path
from datetime import date, timedelta

# results file name and location
yesterday = (date.today() - timedelta(days=1)).strftime('%d.%m.%Y')  # datums līdz
week_ago = (date.today() - timedelta(days=7)).strftime('%d.%m.%Y')
result_folder_path = Path('results')
sample_xlsx = Path('sample_excel.xlsx')   # source excel
xlsx_filename = "laika_apstakli_fakts_prognoze_%s.xlsx" % date.today().strftime('%d%m%Y')
# final_wb = openpyxl.load_workbook('sample_excel.xlsx') # TODO add inside function as var
completename = os.path.join(result_folder_path, xlsx_filename)

# historical weather excel file data folder
station_history = Path('data', date.today().strftime('%Y%m%d'))

# weather stations that should be inserted multiple times in excel file
dupl_st = ['Bauska', 'Gulbene', 'Kuldīga', 'Rūjiena', 'Stende']
tripl_st = ['Zīlāni', 'Zosēni']

station_dict = {
    'Ainaži': 30000, 'Alūksne': 30004, 'Bauska': 30011, 'Daugavpils': 30021, 'Dobele': 30022, 'Gulbene': 30034,
    'Jelgava': 30036, 'Kalnciems': 30040, 'Kolka': 30046, 'Kuldīga': 30048, 'Lielpeči': 30058, 'Liepāja': 30060,
    'Mērsrags': 30072, 'Pāvilosta': 30080, 'Piedruja': 30081, 'Rēzekne': 10000180, 'Rīga': 30096, 'Rūjiena': 30100,
    'Saldus': 30102, 'Sigulda': 30103, 'Sīļi': 30104, 'Skrīveri': 30105, 'Stende': 30111, 'Vičaki': 30132,
    'Zīlāni': 30140, 'Zosēni': 30141
}

# Prepare new xlsx file and update dates in forecast sheet

def create_results_file(sample_xlsx):
    # print('file location: ', sample_file_path)
    wb_obj = openpyxl.load_workbook(sample_xlsx)
    sheet = wb_obj["t_mm_prognoze"]
    dd = 0
    for row in sheet.iter_rows(min_row=3, max_row=12, max_col=5, min_col=5):
        for cell in row:
            val = (date.today() + timedelta(days=dd)).strftime('%d.%m.%Y')
            cell.value = val
            dd += 1
            print(cell.value, end=" ")

    xlsx_filename = "laika_apstakli_fakts_prognoze_%s.xlsx" % date.today().strftime('%d%m%Y')
    completename = os.path.join(result_folder_path, xlsx_filename)
    wb_obj.save(completename)
    return completename

