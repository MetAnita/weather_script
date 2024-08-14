'''''Acquire weather forecast from https://videscentrs.lvgmc.lv'''

import time
import openpyxl
from write_info import completename, dupl_st, tripl_st
import requests


BASE_URL = (
    "https://videscentrs.lvgmc.lv/data/weather_forecast_for_location_daily?punkts={station_point}"
)

# station dict with extracted point values, note Lielpeči = Ogre, Sīļi = Silajāni, Vičaki = Ventspils
station_dict = {'Ainaži': 'P3', 'Alūksne':'P14', 'Bauska': 'P67', 'Daugavpils': 'P75', 'Dobele': 'P54',
                'Gulbene': 'P23', 'Jelgava': 'P52', 'Kalnciems': 'P43', 'Kolka': 'P769', 'Kuldīga': 'P31',
                'Lielpeči': 'P42', 'Liepāja': 'P77', 'Mērsrags': 'P729', 'Pāvilosta': 'P36', 'Piedruja': 'P2388',
                'Rēzekne': 'P62', 'Rīga': 'P28', 'Rūjiena': 'P1', 'Saldus': 'P51', 'Sigulda': 'P24', 'Sīļi': 'P2294',
                'Skrīveri': 'P911', 'Stende': 'P26', 'Vičaki': 'P13','Zīlāni': 'P226','Zosēni': 'P2017'}


def is_float(element) -> bool:
    try:
        float(element)
        return True
    except ValueError:
        return False


# Send a GET request to gather weather forecast JSON file
def get_station_forecast(url):
    response = requests.get(url)
    data = response.json()
    return data


def test_urls(st_dict):
    for station_name, station_point in st_dict.items():
        url = BASE_URL.format(station_point=station_point)
        print(url)


def mm_temp_9_day_forecast(test_data):
    """"get 9 day forecast from lvgmc for single station"""

    day_temperatures = []
    night_temperatures = []
    day_mm = []
    night_mm = []

    # Iterate through the data
    for entry in test_data[1:]:  # skip first entry (today)
        if entry["laiks"].endswith("1200"):
            day_temperatures.append(float(entry["temperatura"]))
            day_mm.append(float(entry["nokrisni_12h"]))
        elif entry["laiks"].endswith("0000"):
            night_temperatures.append(float(entry["temperatura"]))
            night_mm.append(float(entry["nokrisni_12h"]))

    # Print the results
    print("Day Temperatures:", day_temperatures)
    print("Night Temperatures:", night_temperatures)
    print("Day mm:", day_mm)
    print("Night mm:", night_mm)

    return day_temperatures, night_temperatures, day_mm, night_mm


# mm_temp_9_day_forecast(get_station_forecast("https://videscentrs.lvgmc.lv/data/weather_forecast_for_location_daily?punkts=P3"))
# mm_temp_9_day_forecast("Alūksne")


def check_list_len(l):
    """"lvgmc webpage may have incorrect data displayed in some forecasts, this function validates correct lenght
    and prepares data for excel form"""
    if len(l)==9:
        l.append("")
    elif len(l)<9:
        l = [""] * 10
    elif len(l)>9:
        l = [""] * 10
    return l


def mm_temp_forecast_duplicates(station_dict, dupl, tripl):
    """"Modify data by adding duplicate values and empty values to comply with excel structure"""
    stations_day_temp = []
    stations_night_temp = []
    stations_day_mm = []
    stations_night_mm = []
    for station_name, station_point in station_dict.items():
        url = BASE_URL.format(station_point=station_point)
        test_data = get_station_forecast(url)
        print('Getting data for: ', station_name)
        t_day,t_night, mm_day, mm_night = mm_temp_9_day_forecast(test_data)
        t_day = check_list_len(t_day)
        t_night = check_list_len(t_night)
        mm_day = check_list_len(mm_day)
        mm_night = check_list_len(mm_night)
        if station_name in dupl:
            print('Appending 2x data for: ', station_name,' to comply with excel structure' )
            stations_day_temp.extend(2 * t_day)
            stations_night_temp.extend(2 * t_night)
            stations_day_mm.extend(2 * mm_day)
            stations_night_mm.extend(2 * mm_night)
        elif station_name in tripl:
            print('Appending 3x data for: ', station_name, ' to comply with excel structure')
            stations_day_temp.extend(3 * t_day)
            stations_night_temp.extend(3 * t_night)
            stations_day_mm.extend(3 * mm_day)
            stations_night_mm.extend(3 * mm_night)
        else:
            print('Appending data')
            stations_day_temp.extend(t_day)
            stations_night_temp.extend(t_night)
            stations_day_mm.extend(mm_day)
            stations_night_mm.extend(mm_night)
        time.sleep(2)
    return stations_day_temp, stations_night_temp, stations_day_mm, stations_night_mm


def update_forecast_lvgmc(filename):
    wb_obj = openpyxl.load_workbook(filename)
    # Read the forecast sheet t_mm_prognoze
    sheet = wb_obj["t_mm_prognoze"]
    list_day_t, list_night_t, list_day_mm, list_night_mm = mm_temp_forecast_duplicates(station_dict, dupl_st, tripl_st)

    for i in range(3, 353):
        l = i - 3
        cellref_t = sheet.cell(row=i, column=13)     # writing day temp to M column
        cellref_t.value = list_day_t[l]
        cellref_t = sheet.cell(row=i, column=14)     # writing night temp to N column
        cellref_t.value = list_night_t[l]
        cellref_t = sheet.cell(row=i, column=15)     # writing day mm to O column
        cellref_t.value = list_day_mm[l]
        cellref_t = sheet.cell(row=i, column=16)     # writing night mm to P column
        cellref_t.value = list_night_mm[l]


    wb_obj.save(filename)


update_forecast_lvgmc(completename)