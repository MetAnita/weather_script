import openpyxl
from openpyxl.styles import Font
from pathlib import Path
from datetime import date, timedelta
import forecast
import os


xlsx_file = Path('sample_excel.xlsx')   # source excel
result_folder_path = Path('results')

# Save new xlsx file and update dates in forecast sheet

def update_dates(sample_file_path):
    print('file location: ', sample_file_path)
    wb_obj = openpyxl.load_workbook(sample_file_path)
    sheet = wb_obj["t_mm_prognoze"]
    dd = 0
    for row in sheet.iter_rows(min_row=3, max_row=12, max_col=5, min_col=5):
        for cell in row:
            dd += 1
            val = (date.today() + timedelta(days=dd)).strftime('%d.%m.%Y')
            cell.value = val
            # cell.font = Font(color="00000080", italic=True)  # for testing write in blue
            print(cell.value, end=" ")

    xlsx_filename = "laika_apstakli_fakts_prognoze_%s.xlsx" % date.today().strftime('%d%m%Y')
    completename = os.path.join(result_folder_path, xlsx_filename)
    wb_obj.save(completename)
    return completename


def update_forecast(filename):
    wb_obj = openpyxl.load_workbook(filename)
    # Read the forecast sheet t_mm_prognoze
    sheet = wb_obj["t_mm_prognoze"]

    dupl_st = ['Bauska', 'Gulbene', 'Kuldīga', 'Rūjiena', 'Stende']
    tripl_st = ['Zīlāni', 'Zosēni']
    station_dict = {'Ainaži': 'ainazi-47075', 'Alūksne': 'aluksne-4124', 'Bauska': 'bauska-4139', 'Daugavpils': 'daugavpils-4177', 'Dobele': 'dobele-4137', 'Gulbene': 'gulbene-4125', 'Jelgava': 'jelgava-4138', 'Kalnciems': 'kalnciems-322319', 'Kolka': 'kolka-47057', 'Kuldīga': 'kuldiga-4106', 'Lielpeči': 'ogre-47066', 'Liepāja': 'liepaja-4134', 'Mērsrags': 'mersrags-47065', 'Pāvilosta': 'pavilosta-47067', 'Piedruja': 'piedruja-322361', 'Rēzekne': 'rezekne-4140', 'Rīga': 'riga-4136', 'Rūjiena': 'rujiena-47073', 'Saldus': 'saldus-4135', 'Sigulda': 'sigulda-4103', 'Sīļi': 'silajani-322513', 'Skrīveri': 'skriveri-322563', 'Stende': 'stende-47083', 'Vičaki': 'ventspils-4123', 'Zīlāni': 'zilani-322369', 'Zosēni': 'vecpiebalga-322623'}

    listt, listmm = forecast.mm_temp_forecast_duplicates(station_dict, dupl_st, tripl_st)
    # l = 0           # iterable for temperature list

    for i in range(3, 353):
        l = i - 3
        cellref_t = sheet.cell(row=i, column=6)     # writing temp to f column
        cellref_t.value = listt[l]
        cellref_mm = sheet.cell(row=i, column=8)    # writing mm to H column
        cellref_mm.value = listmm[l]

    wb_obj.save(filename)


new_file = update_dates(xlsx_file)
update_forecast(new_file)
