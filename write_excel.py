''''Write historical mm and t data from meteo to excel file'''
from pathlib import Path
import openpyxl
import xlrd
from write_info import completename, station_history, sample_data_filepath

station_dict = {'Ainaži': 'ainazi-47075', 'Alūksne': 'aluksne-4124', 'Bauska': 'bauska-4139', 'Daugavpils': 'daugavpils-4177', 'Dobele': 'dobele-4137', 'Gulbene': 'gulbene-4125', 'Jelgava': 'jelgava-4138', 'Kalnciems': 'kalnciems-322319', 'Kolka': 'kolka-47057', 'Kuldīga': 'kuldiga-4106', 'Lielpeči': 'ogre-47066', 'Liepāja': 'liepaja-4134', 'Mērsrags': 'mersrags-47065', 'Pāvilosta': 'pavilosta-47067', 'Piedruja': 'piedruja-322361', 'Rēzekne': 'rezekne-4140', 'Rīga': 'riga-4136', 'Rūjiena': 'rujiena-47073', 'Saldus': 'saldus-4135', 'Sigulda': 'sigulda-4103', 'Sīļi': 'silajani-322513', 'Skrīveri': 'skriveri-322563', 'Stende': 'stende-47083', 'Vičaki': 'ventspils-4123', 'Zīlāni': 'zilani-322369', 'Zosēni': 'vecpiebalga-322623'}
final_wb = openpyxl.load_workbook(completename)
final_t_sheet = final_wb['t_fakts']


# find station row in excel final file
def find_start_row(station):
    row_idx = False
    for row in final_t_sheet.iter_rows(min_col=2, max_col=2):   # B column in Excel file
        for cell in row:
            if cell.value == station:
                # print('Row number:', row[0].row)
                row_idx = row[0].row
    return row_idx


def write_data_chunk(st_row, src_ws, dest_ws):
    h = 4
    for i in range(st_row, st_row+7):  # destination excel rows from - to
        # get each row from source file into list
        #header = src_ws.row_values(h, start_colx=0, end_colx=None)
        header = [cell.value for cell in src_ws[h]]  # h is the row index
        print(header)
        h += 1
        if st_row == 3:  # write dates & data for first station
            for j in range(4, 29):  # columns (with date column)
                # reading cell value from source excel file
                c = header[j - 4]
                # writing the read value to destination excel file
                dest_ws.cell(row=i, column=j).value = c
        else:
            for j in range(5, 29):  # columns (without date column)
                # reading cell value from source excel file
                c = header[j - 4]
                # writing the read value to destination excel file
                dest_ws.cell(row=i, column=j).value = c


def write_t_history(src_filepath, dest_filepath, station):
    station_wb = openpyxl.load_workbook(src_filepath)
    final_wb = openpyxl.load_workbook(dest_filepath)
    t_ws = station_wb.worksheets[0]
    final_t_ws = final_wb['t_fakts']
    str_row = find_start_row(station)  # place where to write data in dest worksheet
    # write temperature
    write_data_chunk(str_row, t_ws, final_t_ws)
    final_wb.save(dest_filepath)


def write_mm_history(src_filepath, dest_filepath, station):
    station_wb = openpyxl.load_workbook(src_filepath)
    final_wb = openpyxl.load_workbook(dest_filepath)
    t_ws = station_wb.worksheets[0]
    final_mm_ws = final_wb['mm_fakts']
    str_row = find_start_row(station)   # place where to write data in dest worksheet
    # write mm
    write_data_chunk(str_row, t_ws, final_mm_ws)
    final_wb.save(dest_filepath)


for key in station_dict:
    print(key)
    # write temperature
    st_path = Path(station_history, '%s_temp.xlsx' % key)
    write_t_history(st_path, completename, key)
    # write mm
    st_path = Path(station_history, '%s_mm.xlsx' % key)
    write_mm_history(st_path, completename, key)
