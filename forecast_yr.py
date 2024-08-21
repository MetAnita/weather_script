"""Acquire weather forecast from yr.no and write in results excel file"""
from bs4 import BeautifulSoup
import requests
import time
import re
import openpyxl
from write_info import completename, dupl_st, tripl_st


STATION_DICT = {'Ainaži': '2-461628', 'Alūksne': '2-461528', 'Bauska': '2-461114', 'Daugavpils': '2-460413',
                'Dobele': '2-460312', 'Gulbene': '2-459668', 'Jelgava': '2-459279', 'Kalnciems': '2-459102',
                'Kolka': '2-458682', 'Kuldīga': '2-458460', 'Lielpeči': '2-457065', 'Liepāja': '5-2640600',
                'Mērsrags': '2-457408', 'Pāvilosta': '2-456827', 'Piedruja': '2-456742', 'Rēzekne': '2-456202',
                'Rīga': '2-456172', 'Rūjiena': '2-456008', 'Saldus': '2-455890', 'Sigulda': '2-455718',
                'Sīļi': '2-455697', 'Skrīveri': '2-455523', 'Stende': '2-455260', 'Vičaki': '2-454251',
                'Zīlāni': '2-453862', 'Zosēni': '2-453822'}

HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET',
    'Access-Control-Allow-Headers': 'Content-Type',
    'Access-Control-Max-Age': '3600',
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0'
    }


def fetch_weather_data(station):
    """Fetch 9-day weather forecast for a single station."""
    station_code = STATION_DICT.get(station)
    if not station_code:
        raise ValueError(f"Station '{station}' not found in station dictionary.")

    url = f"https://www.yr.no/en/forecast/daily-table/{station_code}"
    print(f"Fetching data from: {url}")

    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Error fetching data for {station}: {e}")
        return None, None

    soup = BeautifulSoup(response.content, "lxml")
    return parse_weather_data(soup)


def parse_weather_data(soup):
    """Parse temperature and precipitation data from the HTML soup."""
    min_temps = extract_temperatures(soup, 'min')
    max_temps = extract_temperatures(soup, 'max')
    precipitations = extract_precipitations(soup)

    avg_temps = [(int(min_t) + int(max_t)) / 2 for min_t, max_t in zip(min_temps, max_temps)]

    return avg_temps, precipitations


def extract_temperatures(soup, temp_type):
    """Extract min or max temperatures based on the type."""
    class_prefix = 'min-max-temperature__' + temp_type
    temperature_spans = soup.find_all('span', {'class': re.compile(rf'^temperature {class_prefix} temperature')})

    def clean_temp(temp_str):
        """Remove non-numeric characters and convert to int."""
        temp_str = temp_str.replace('°', '').strip()  # Remove the degree symbol and any extra whitespace
        return int(temp_str)

    return [clean_temp(span.get_text()) for span in temperature_spans]     #  temperature_spans[1:]]  # skip today


def extract_precipitations(soup):
    """Extract precipitation data from the soup."""
    precipitations = []
    precip_divs = soup.find_all('div', class_='daily-weather-list-item__precipitation')  # [1:]  # skip today

    for div in precip_divs:
        span = div.find('span', class_='Precipitation-module__main-sU6qN')
        if span:
            value_span = span.find_all('span')[1]
            value = float(value_span.get_text(strip=True).replace(',', '.'))
            precipitations.append(value)

    return precipitations


def compile_forecast_data():
    """Compile weather data for all stations."""
    temperatures, precipitations = [], []

    for station in STATION_DICT:
        print(f"Processing {station}...")
        avg_temps, mm = fetch_weather_data(station)
        if avg_temps and mm:
            #avg_temps.append("")  # Placeholder for Excel formatting
            #mm.append("")
            if station in dupl_st:
                temperatures.extend(2 * avg_temps)
                precipitations.extend(2 * mm)
            elif station in tripl_st:
                temperatures.extend(3 * avg_temps)
                precipitations.extend(3 * mm)
            else:
                temperatures.extend(avg_temps)
                precipitations.extend(mm)
        time.sleep(1)

    return temperatures, precipitations


def update_excel_with_forecast(filename):
    """Update the Excel file with the weather forecast."""
    wb = openpyxl.load_workbook(filename)
    sheet = wb["t_mm_prognoze"]

    temperatures, precipitations = compile_forecast_data()

    for i in range(3, 353):
        sheet.cell(row=i, column=10, value=temperatures[i - 3])         # writing temp to J column
        sheet.cell(row=i, column=11, value=precipitations[i - 3])       # writing mm to K column

    wb.save(filename)


if __name__ == "__main__":
   update_excel_with_forecast(completename)

